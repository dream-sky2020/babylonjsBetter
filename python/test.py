import tkinter as tk
from tkinter import filedialog, messagebox
import openpyxl
from openpyxl.drawing.image import Image as OpenpyxlImage
import qrcode
import io
import os

def process_excel():
    # 1. 初始化 tkinter 并隐藏主窗口
    root = tk.Tk()
    root.withdraw()
    
    # 2. 弹出文件选择对话框
    file_path = filedialog.askopenfilename(
        title="请选择需要处理的 Excel 文件",
        filetypes=[("Excel 文件", "*.xlsx")]
    )

    if not file_path:
        return

    try:
        # 加载 Excel 文件
        wb = openpyxl.load_workbook(file_path)
        target_col_name = "料号"
        
        # 记录成功处理了多少个 Sheet 页
        sheets_processed = 0

        # 3. 遍历工作簿中的每一个 Sheet
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            target_col_idx = -1
            header_row_idx = -1

            # 在当前 Sheet 扫描前 10 行寻找 "料号"
            for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), 1):
                for c_idx, cell_value in enumerate(row, 1):
                    if str(cell_value).strip() == target_col_name:
                        target_col_idx = c_idx
                        header_row_idx = r_idx
                        break
                if target_col_idx != -1:
                    break

            # 如果当前 Sheet 没有找到 "料号" 表头，直接跳过，检查下一个 Sheet
            if target_col_idx == -1:
                continue
            
            # 如果找到了，计数器加1，开始处理当前 Sheet
            sheets_processed += 1

            # 4. 插入新列
            new_col_idx = target_col_idx + 1
            ws.insert_cols(new_col_idx)
            ws.cell(row=header_row_idx, column=new_col_idx, value="料号(二维码)")

            new_col_letter = openpyxl.utils.get_column_letter(new_col_idx)
            ws.column_dimensions[new_col_letter].width = 15

            # 5. 遍历数据生成二维码
            for row in range(header_row_idx + 1, ws.max_row + 1):
                part_no_value = ws.cell(row=row, column=target_col_idx).value

                if part_no_value is not None and str(part_no_value).strip() != "":
                    part_no_str = str(part_no_value).strip()

                    # 生成二维码
                    qr = qrcode.QRCode(
                        version=1,
                        error_correction=qrcode.constants.ERROR_CORRECT_L,
                        box_size=3,
                        border=1,
                    )
                    qr.add_data(part_no_str)
                    qr.make(fit=True)
                    img = qr.make_image(fill_color="black", back_color="white")

                    # 转为字节流
                    img_byte_arr = io.BytesIO()
                    img.save(img_byte_arr, format='PNG')
                    img_byte_arr.seek(0)

                    # 写入 Excel
                    xl_img = OpenpyxlImage(img_byte_arr)
                    xl_img.width = 80
                    xl_img.height = 80

                    ws.row_dimensions[row].height = 65
                    cell_coord = f"{new_col_letter}{row}"
                    ws.add_image(xl_img, cell_coord)

        # 6. 检查是否所有的 Sheet 都没有找到料号
        if sheets_processed == 0:
            messagebox.showwarning("提示", f"在文件的任何子表中都未找到名为 '{target_col_name}' 的表头！")
            return

        # 7. 保存文件
        dir_name = os.path.dirname(file_path)
        base_name = os.path.basename(file_path)
        name, ext = os.path.splitext(base_name)
        new_file_path = os.path.join(dir_name, f"{name}_已生成二维码{ext}")

        wb.save(new_file_path)
        messagebox.showinfo("成功", f"处理完成！共为您处理了 {sheets_processed} 个子表。\n文件已保存至：\n{new_file_path}")

    except Exception as e:
        messagebox.showerror("错误", f"处理过程中发生错误：\n{str(e)}")

if __name__ == "__main__":
    process_excel()